"""Multi-sheet Excel workbook from report queries."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from sqlalchemy.orm import Session

from harness_analytics.reports import (
    analytics_column_header,
    report_all_applications,
    report_all_harness,
    report_art_unit_summary,
    report_art_unit_summary_all_applications,
    report_by_office,
    report_by_office_all_applications,
    report_interview_to_noa,
    report_interview_to_noa_all_applications,
    report_specific_clients,
    report_specific_clients_all_applications,
)

JAC_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
YEAR_FILLS = {
    2024: PatternFill("solid", fgColor="EBF3FB"),
    2025: PatternFill("solid", fgColor="E2EFDA"),
}


def _office_sheet_label(office_name: object) -> str:
    if office_name is None:
        return "UNKNOWN"
    try:
        if pd.isna(office_name):
            return "UNKNOWN"
    except (TypeError, ValueError):
        pass
    s = str(office_name).strip()
    return (s or "UNKNOWN")[:31]


def _write_df_to_sheet(ws, df: pd.DataFrame, *, highlight_jac: bool = False) -> None:
    if df.empty:
        ws.cell(row=1, column=1, value="No rows")
        return

    cols = list(df.columns)
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=analytics_column_header(str(col_name)))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        issue_year = getattr(row, "issue_year", None) if hasattr(row, "issue_year") else None
        is_jac = bool(getattr(row, "is_jac", False)) if hasattr(row, "is_jac") else False

        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if highlight_jac and is_jac:
                cell.fill = JAC_FILL
            elif issue_year in YEAR_FILLS:
                cell.fill = YEAR_FILLS[int(issue_year)]

    for col_idx, col_name in enumerate(cols, 1):
        max_len = max(len(str(col_name)), 10, *(len(str(c)) for c in df[col_name].head(200)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _write_summary_tab(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws.cell(row=1, column=1, value="No data for summary")
        return

    rows = [
        ["Metric", "2024", "2025"],
        [
            "Total Applications Issued",
            int((df.issue_year == 2024).sum()),
            int((df.issue_year == 2025).sum()),
        ],
        [
            "Avg Substantive OAs Before NOA",
            round(df.loc[df.issue_year == 2024, "total_substantive_oas"].mean(), 2),
            round(df.loc[df.issue_year == 2025, "total_substantive_oas"].mean(), 2),
        ],
        [
            "Avg Non-Final OAs",
            round(df.loc[df.issue_year == 2024, "nonfinal_oa_count"].mean(), 2),
            round(df.loc[df.issue_year == 2025, "nonfinal_oa_count"].mean(), 2),
        ],
        [
            "Avg Final OAs",
            round(df.loc[df.issue_year == 2024, "final_oa_count"].mean(), 2),
            round(df.loc[df.issue_year == 2025, "final_oa_count"].mean(), 2),
        ],
        [
            "% With 0 OAs (straight to NOA)",
            f"{100 * (df.loc[df.issue_year == 2024, 'total_substantive_oas'] == 0).mean():.1f}%",
            f"{100 * (df.loc[df.issue_year == 2025, 'total_substantive_oas'] == 0).mean():.1f}%",
        ],
        [
            "% With ≥ 1 Final OA",
            f"{100 * (df.loc[df.issue_year == 2024, 'final_oa_count'] >= 1).mean():.1f}%",
            f"{100 * (df.loc[df.issue_year == 2025, 'final_oa_count'] >= 1).mean():.1f}%",
        ],
        [
            "Avg Days Filing to NOA",
            round(df.loc[df.issue_year == 2024, "days_filing_to_noa"].mean(), 0),
            round(df.loc[df.issue_year == 2025, "days_filing_to_noa"].mean(), 0),
        ],
        [
            "Applications with Interviews",
            int((df.loc[df.issue_year == 2024, "had_examiner_interview"]).sum()),
            int((df.loc[df.issue_year == 2025, "had_examiner_interview"]).sum()),
        ],
        [
            "NOA within interview window of last interview",
            int((df.loc[df.issue_year == 2024, "interview_led_to_noa"]).sum()),
            int((df.loc[df.issue_year == 2025, "interview_led_to_noa"]).sum()),
        ],
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL


def _write_summary_tab_multi_year(ws, df: pd.DataFrame) -> None:
    """Like _write_summary_tab but one column per distinct issue_year in the frame."""
    if df.empty or "issue_year" not in df.columns:
        ws.cell(row=1, column=1, value="No data for summary")
        return
    year_norm = pd.to_numeric(df["issue_year"], errors="coerce").dropna().astype(int)
    if year_norm.empty:
        ws.cell(row=1, column=1, value="No issue_year values for summary")
        return
    years_list = sorted(year_norm.unique().tolist())

    def _mask_year(y: int) -> pd.Series:
        return pd.to_numeric(df["issue_year"], errors="coerce") == y

    def _count(y: int) -> int:
        return int(_mask_year(y).sum())

    def _pct_zero_oa(y: int) -> str:
        sub = df.loc[_mask_year(y), "total_substantive_oas"]
        if sub.empty:
            return ""
        m = (sub == 0).mean()
        if pd.isna(m):
            return ""
        return f"{100 * float(m):.1f}%"

    def _pct_final(y: int) -> str:
        sub = df.loc[_mask_year(y), "final_oa_count"]
        if sub.empty:
            return ""
        m = (sub >= 1).mean()
        if pd.isna(m):
            return ""
        return f"{100 * float(m):.1f}%"

    def _fmt_mean_masked(col: str, year: int, nd: int) -> object:
        sub = df.loc[_mask_year(year), col]
        if sub.empty:
            return None
        m = sub.mean()
        if pd.isna(m):
            return None
        return round(float(m), nd)

    header = ["Metric"] + [str(y) for y in years_list]
    rows: list[list[object]] = [
        header,
        ["Applications (by issue year)"] + [_count(y) for y in years_list],
        ["Avg Substantive OAs Before NOA"] + [_fmt_mean_masked("total_substantive_oas", y, 2) for y in years_list],
        ["Avg Non-Final OAs"] + [_fmt_mean_masked("nonfinal_oa_count", y, 2) for y in years_list],
        ["Avg Final OAs"] + [_fmt_mean_masked("final_oa_count", y, 2) for y in years_list],
        ["% With 0 OAs (straight to NOA)"] + [_pct_zero_oa(y) for y in years_list],
        ["% With ≥ 1 Final OA"] + [_pct_final(y) for y in years_list],
        ["Avg Days Filing to NOA"] + [_fmt_mean_masked("days_filing_to_noa", y, 0) for y in years_list],
        [
            "Applications with Interviews",
        ]
        + [int((df.loc[_mask_year(y), "had_examiner_interview"]).sum()) for y in years_list],
        [
            "NOA within interview window of last interview",
        ]
        + [int((df.loc[_mask_year(y), "interview_led_to_noa"]).sum()) for y in years_list],
    ]
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL


def build_excel_workbook(db: Session) -> Workbook:
    """Build the full multi-tab Excel workbook in memory (same sheets as CLI report)."""
    wb = Workbook()
    wb.remove(wb.active)

    df_all = report_all_harness(db)
    ws1 = wb.create_sheet("All Harness IP")
    _write_df_to_sheet(ws1, df_all, highlight_jac=True)

    offices = report_by_office(db)
    for office_name, df_office in offices.items():
        safe = _office_sheet_label(office_name)
        ws = wb.create_sheet(f"Office - {safe}")
        _write_df_to_sheet(ws, df_office, highlight_jac=(office_name == "DC"))

    df_clients = report_specific_clients(db, ["15639", "2557SI", "9587SI"])
    ws3 = wb.create_sheet("Samsung Clients")
    _write_df_to_sheet(ws3, df_clients, highlight_jac=True)

    df_int = report_interview_to_noa(db)
    ws4 = wb.create_sheet("Interview to NOA")
    _write_df_to_sheet(ws4, df_int)

    df_au = report_art_unit_summary(db)
    ws5 = wb.create_sheet("Art Unit Summary")
    _write_df_to_sheet(ws5, df_au)

    ws6 = wb.create_sheet("Summary Statistics")
    _write_summary_tab(ws6, df_all)

    return wb


def build_excel_workbook_all_applications(db: Session) -> Workbook:
    """Multi-tab workbook: every application (any status), analytics columns when present."""
    wb = Workbook()
    wb.remove(wb.active)

    df_all = report_all_applications(db)
    ws1 = wb.create_sheet("All applications")
    _write_df_to_sheet(ws1, df_all, highlight_jac=True)

    offices = report_by_office_all_applications(db)
    for office_name, df_office in offices.items():
        safe = _office_sheet_label(office_name)
        ws = wb.create_sheet(f"Office - {safe}")
        _write_df_to_sheet(ws, df_office, highlight_jac=(str(office_name) == "DC"))

    df_clients = report_specific_clients_all_applications(db, ["15639", "2557SI", "9587SI"])
    ws3 = wb.create_sheet("Samsung Clients")
    _write_df_to_sheet(ws3, df_clients, highlight_jac=True)

    df_int = report_interview_to_noa_all_applications(db)
    ws4 = wb.create_sheet("Interview to NOA")
    _write_df_to_sheet(ws4, df_int)

    df_au = report_art_unit_summary_all_applications(db)
    ws5 = wb.create_sheet("Art Unit Summary")
    _write_df_to_sheet(ws5, df_au)

    ws6 = wb.create_sheet("Summary Statistics")
    _write_summary_tab_multi_year(ws6, df_all)

    return wb


def workbook_to_bytesio(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_excel_report(db: Session, output_path: str) -> None:
    """Build the full multi-tab Excel workbook and save to disk."""
    wb = build_excel_workbook(db)
    wb.save(output_path)
